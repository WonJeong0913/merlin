"""Run bounded C0/C1 engineering pilots through the scripted-solver harness.

This is a Merlin-specific engineering harness, not a paper-faithful
SkillsBench C0/C1 implementation. The model does not freely operate on the
workspace and C1 skill content may be prompt-injected under a size bound.
Each trial follows this protocol:

1. build the task image,
2. start a fresh task container,
3. inspect the container state,
4. ask the model to generate one solve script,
5. run the generated script inside the container,
6. run the upstream verifier,
7. optionally regenerate once from the script/verifier failure signal.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shlex
import shutil
import subprocess
import sys
import tempfile
import time
import uuid
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent
REPO_ROOT = ROOT.parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from experiments.skillsbench.run_model_c0_c1_pilot import (
    DEFAULT_MATRIX,
    DEFAULT_ORACLE_SUMMARY,
    extract_account_usage,
    load_conditions,
    ready_task_ids,
)
from experiments.skillsbench.run_oracle_readiness import (
    CommandReport,
    classify_verifier_result,
    docker_resource_args,
    ensure_docker,
    prepare_skill_free_build_context,
    read_reward,
    run_command,
    safe_name,
    stop_process,
    tail_text,
)


TASKS_ROOT = ROOT / "tasks"
DEFAULT_SPLIT_MANIFEST = ROOT / "split-manifest.json"
DEFAULT_RUNS_ROOT = ROOT / "runs" / "model-c0-c1-scripted"
DEFAULT_TASKS = ["weighted-gdp-calc"]
TAIL_CHARS = 12000
MAX_INSPECTION_CHARS = 12000
MAX_SKILL_CHARS = 24000
GENERATION_TRACE_CHARS = 12000
MAX_GENERATED_SCRIPT_CHARS = 20000
MAX_GENERATION_NOTES_CHARS = 1000
RESPONSE_CONTRACT_VERSION = "script-json-v2"

SCRIPT_RESPONSE_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "script_path": {
            "type": "string",
            "enum": ["solve.sh", "solution.py"],
        },
        "script": {
            "type": "string",
            "minLength": 1,
            "maxLength": MAX_GENERATED_SCRIPT_CHARS,
        },
        "skill_used": {
            "type": "array",
            "items": {"type": "string", "maxLength": 512},
            "maxItems": 32,
        },
        "notes": {
            "type": "string",
            "maxLength": MAX_GENERATION_NOTES_CHARS,
        },
    },
    "required": ["script_path", "script", "skill_used", "notes"],
    "additionalProperties": False,
}


@dataclass(slots=True)
class ScriptedRecord:
    task_id: str
    condition_id: str
    arm: str
    harness_mode: str
    model_id: str
    backend: str
    effort: str
    runtime_effort: str
    status: str
    passed: bool
    response_contract_version: str = RESPONSE_CONTRACT_VERSION
    backend_type: str = "B_cli"
    auth_mode: str = "user_owned_account"
    credential_forwarded_to_container: bool = False
    trial_index: int = 1
    trial_id: str | None = None
    seed_control: str = "provider_cli_unavailable"
    temperature_control: str = "provider_default_or_unset"
    reward: float | None = None
    wall_time_sec: float | None = None
    container_workdir: str | None = None
    script_path: str | None = None
    skill_used: list[str] = field(default_factory=list)
    skill_usage_evidence: str = "model_self_report_only"
    task_instruction_sha256: str | None = None
    prompt_sha256: str | None = None
    skill_context_report: dict[str, Any] = field(default_factory=dict)
    repair_iterations_used: int = 0
    account_usage: dict[str, Any] = field(default_factory=dict)
    timeouts: dict[str, int] = field(default_factory=dict)
    logs_dir: str | None = None
    commands: dict[str, CommandReport] = field(default_factory=dict)
    notes: list[str] = field(default_factory=list)


def run_command_with_tail(argv: list[str], *, cwd: Path | None = None, timeout_sec: int) -> CommandReport:
    report = run_command(argv, cwd=cwd, timeout_sec=timeout_sec)
    report.stdout_tail = tail_text(report.stdout_tail, TAIL_CHARS)
    report.stderr_tail = tail_text(report.stderr_tail, TAIL_CHARS)
    return report


def normalize_container_workdir(raw: str | None) -> str:
    value = (raw or "").strip()
    if not value:
        return "/root"
    if not value.startswith("/"):
        return f"/{value}"
    return value


def detect_image_workdir(image: str) -> tuple[str, CommandReport]:
    report = run_command_with_tail(
        ["docker", "image", "inspect", "-f", "{{.Config.WorkingDir}}", image],
        timeout_sec=30,
    )
    if report.exit_code != 0:
        return "/root", report
    return normalize_container_workdir(report.stdout_tail), report


def read_task_prompt(task_dir: Path) -> str:
    return (task_dir / "task.md").read_text(encoding="utf-8", errors="replace")


def task_instruction_body(task_text: str) -> str:
    """Return only the user-visible task body, withholding benchmark metadata."""

    if not task_text.startswith("---"):
        return task_text
    parts = task_text.split("---", 2)
    if len(parts) < 3:
        return task_text
    return parts[2].lstrip("\r\n")


def frontmatter(task_text: str) -> str:
    if not task_text.startswith("---"):
        return ""
    parts = task_text.split("---", 2)
    return parts[1] if len(parts) >= 3 else ""


def section_timeout(task_text: str, section: str, default: int) -> int:
    fm = frontmatter(task_text)
    match = re.search(rf"(?ms)^{re.escape(section)}:\n(?P<body>(?:^[ \t]+.*\n?)*)", fm)
    if not match:
        return default
    timeout = re.search(r"(?m)^[ \t]+timeout_sec:\s*([0-9.]+)", match.group("body"))
    if not timeout:
        return default
    return max(1, int(float(timeout.group(1))))


def task_timeouts(task_text: str, *, build_default: int, script_default: int, verifier_default: int) -> dict[str, int]:
    return {
        "build_timeout_sec": section_timeout_field(
            task_text,
            "environment",
            "build_timeout_sec",
            build_default,
        ),
        "script_execution_timeout_sec": section_timeout(task_text, "agent", script_default),
        "verifier_timeout_sec": section_timeout(task_text, "verifier", verifier_default),
    }


def section_timeout_field(task_text: str, section: str, field: str, default: int) -> int:
    fm = frontmatter(task_text)
    match = re.search(rf"(?ms)^{re.escape(section)}:\n(?P<body>(?:^[ \t]+.*\n?)*)", fm)
    if not match:
        return default
    timeout = re.search(
        rf"(?m)^[ \t]+{re.escape(field)}:\s*([0-9.]+)",
        match.group("body"),
    )
    if not timeout:
        return default
    return max(1, int(float(timeout.group(1))))


def load_adaptation_ready_tasks(split_path: Path, oracle_summary: Path) -> list[str]:
    split = json.loads(split_path.read_text(encoding="utf-8"))
    ready = ready_task_ids(oracle_summary)
    return [
        item["task_id"]
        for item in split["splits"]["adaptation"]
        if item["task_id"] in ready
    ]


def collect_skill_context(task_dir: Path, *, arm: str, max_chars: int = MAX_SKILL_CHARS) -> list[dict[str, str]]:
    if arm != "C1":
        return []
    skills_dir = task_dir / "environment" / "skills"
    if not skills_dir.exists():
        return []
    remaining = max_chars
    items: list[dict[str, str]] = []
    text_suffixes = {".md", ".txt", ".py", ".js", ".ts", ".json", ".yaml", ".yml", ".skill", ".sh"}
    for path in sorted(item for item in skills_dir.rglob("*") if item.is_file()):
        if remaining <= 0:
            break
        if path.suffix.lower() not in text_suffixes:
            continue
        try:
            content = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            continue
        snippet = content[:remaining]
        remaining -= len(snippet)
        items.append(
            {
                "path": path.relative_to(skills_dir).as_posix(),
                "content": snippet,
            }
        )
    return items


def describe_skill_context(
    task_dir: Path,
    *,
    arm: str,
    injected: list[dict[str, str]],
    max_chars: int = MAX_SKILL_CHARS,
) -> dict[str, Any]:
    skills_dir = task_dir / "environment" / "skills"
    source_files = sorted(path for path in skills_dir.rglob("*") if path.is_file()) if skills_dir.exists() else []
    text_suffixes = {".md", ".txt", ".py", ".js", ".ts", ".json", ".yaml", ".yml", ".skill", ".sh"}
    text_files = [path for path in source_files if path.suffix.lower() in text_suffixes]
    total_text_chars = 0
    for path in text_files:
        try:
            total_text_chars += len(path.read_text(encoding="utf-8"))
        except UnicodeDecodeError:
            continue
    injected_chars = sum(len(item["content"]) for item in injected)
    source_hasher = hashlib.sha256()
    for path in source_files:
        relative = path.relative_to(skills_dir).as_posix().encode("utf-8")
        source_hasher.update(len(relative).to_bytes(8, "big"))
        source_hasher.update(relative)
        content = path.read_bytes()
        source_hasher.update(len(content).to_bytes(8, "big"))
        source_hasher.update(content)
    injected_payload = json.dumps(injected, sort_keys=True, ensure_ascii=False).encode("utf-8")
    return {
        "mode": "none" if arm == "C0" else "prompt_injected_truncated",
        "source_file_count": len(source_files),
        "eligible_text_file_count": len(text_files),
        "injected_file_count": len(injected),
        "total_text_chars": total_text_chars,
        "injected_chars": injected_chars,
        "max_chars": max_chars,
        "truncated": arm == "C1" and (len(injected) < len(text_files) or injected_chars < total_text_chars),
        "binary_or_unsupported_files_omitted": max(0, len(source_files) - len(text_files)),
        "source_bundle_sha256": source_hasher.hexdigest(),
        "injected_context_sha256": hashlib.sha256(injected_payload).hexdigest(),
    }


def probe_backend_version(condition: dict[str, Any]) -> CommandReport:
    executable = "claude" if condition["backend"] == "claude" else "codex"
    return run_command_with_tail([executable, "--version"], timeout_sec=30)


def make_agent_command(condition: dict[str, Any]) -> list[str]:
    backend = condition["backend"]
    model = condition["model_id"]
    runtime_effort = condition.get("runtime_effort", condition["effort"])
    if backend == "claude":
        return [
            "claude",
            "-p",
            "--output-format",
            "json",
            "--json-schema",
            json.dumps(SCRIPT_RESPONSE_SCHEMA, separators=(",", ":")),
            "--no-session-persistence",
            "--tools",
            "",
            "--model",
            model,
            "--effort",
            runtime_effort,
        ]
    if backend == "codex":
        return [
            "codex",
            "exec",
            "-c",
            f'model_reasoning_effort="{runtime_effort}"',
            "--skip-git-repo-check",
            "--sandbox",
            "read-only",
            "--ignore-rules",
            "--ephemeral",
            "--model",
            model,
            "-",
        ]
    raise ValueError(f"Unsupported backend: {backend}")


def call_generation_model(
    condition: dict[str, Any],
    prompt: str,
    *,
    cwd: Path,
    timeout_sec: int,
) -> CommandReport:
    command = make_agent_command(condition)
    start = time.monotonic()
    try:
        completed = subprocess.run(
            command,
            input=prompt,
            cwd=cwd,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=timeout_sec,
        )
        return CommandReport(
            argv=command,
            exit_code=completed.returncode,
            duration_sec=round(time.monotonic() - start, 3),
            stdout_tail=completed.stdout,
            stderr_tail=tail_text(completed.stderr, TAIL_CHARS),
        )
    except subprocess.TimeoutExpired as exc:
        return CommandReport(
            argv=command,
            exit_code=124,
            duration_sec=round(time.monotonic() - start, 3),
            timed_out=True,
            stdout_tail=exc.stdout or "",
            stderr_tail=tail_text(exc.stderr or "", TAIL_CHARS),
        )


def trim_generation_trace(report: CommandReport) -> None:
    report.stdout_tail = tail_text(report.stdout_tail, GENERATION_TRACE_CHARS)
    report.stderr_tail = tail_text(report.stderr_tail, TAIL_CHARS)


def unwrap_model_text(stdout: str) -> str:
    stripped = stdout.strip()
    try:
        wrapper = json.loads(stripped)
    except json.JSONDecodeError:
        return stripped
    if isinstance(wrapper, dict):
        structured_output = wrapper.get("structured_output")
        if isinstance(structured_output, dict):
            return json.dumps(structured_output, ensure_ascii=False)
        for key in ("result", "output_text", "text", "message"):
            value = wrapper.get(key)
            if isinstance(value, str):
                return value.strip()
        content = wrapper.get("content")
        if isinstance(content, list):
            parts: list[str] = []
            for item in content:
                if isinstance(item, dict) and isinstance(item.get("text"), str):
                    parts.append(item["text"])
            if parts:
                return "\n".join(parts).strip()
    return stripped


def generation_parse_failure_status(stdout: str) -> str:
    """Separate provider/output truncation from ordinary JSON contract failures."""

    stripped = stdout.strip()
    try:
        wrapper = json.loads(stripped)
    except json.JSONDecodeError:
        wrapper = None
    if not isinstance(wrapper, dict):
        return "generation_parse_failed"

    if isinstance(wrapper.get("structured_output"), dict):
        return "generation_parse_failed"

    result = wrapper.get("result")
    if not isinstance(result, str):
        return "generation_parse_failed"
    normalized = result.lower()
    truncation_markers = (
        "output length limit",
        "output limit",
        "maximum output",
        "max output",
        "truncated",
        "forcefully cut off",
        "截断",
        "输出长度限制",
    )
    if any(marker in normalized for marker in truncation_markers):
        return "generation_output_truncated"
    return "generation_parse_failed"


def verifier_outcome_status(report: CommandReport, reward: float | None) -> tuple[str, bool]:
    """Classify task failure separately from verifier infrastructure failure."""

    return classify_verifier_result(report, reward)


def extract_json_object(text: str) -> dict[str, Any]:
    stripped = text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```(?:json)?\s*", "", stripped)
        stripped = re.sub(r"\s*```$", "", stripped)
    try:
        parsed = json.loads(stripped)
    except json.JSONDecodeError:
        start = stripped.find("{")
        end = stripped.rfind("}")
        if start < 0 or end <= start:
            raise ValueError("Model output did not contain a JSON object.")
        parsed = json.loads(stripped[start : end + 1])
    if not isinstance(parsed, dict):
        raise ValueError("Model output JSON must be an object.")
    return parsed


def parse_generated_script(report: CommandReport) -> tuple[str, str, list[str], str]:
    text = unwrap_model_text(report.stdout_tail)
    parsed = extract_json_object(text)
    script = parsed.get("script")
    if not isinstance(script, str) or not script.strip():
        raise ValueError("Generated JSON must include non-empty string field 'script'.")
    if len(script) > MAX_GENERATED_SCRIPT_CHARS:
        raise ValueError(
            f"Generated script exceeds {MAX_GENERATED_SCRIPT_CHARS} characters: {len(script)}."
        )
    script_path = parsed.get("script_path")
    if not isinstance(script_path, str) or not script_path.strip():
        raise ValueError("Generated JSON must include non-empty string field 'script_path'.")
    script_name = Path(script_path).name
    if script_name not in {"solve.sh", "solution.py"}:
        raise ValueError(f"Unsupported script_path: {script_path}")
    if not re.fullmatch(r"[A-Za-z0-9_.-]+", script_name):
        raise ValueError(f"Unsafe script_path: {script_path}")
    skill_used_raw = parsed.get("skill_used")
    if not isinstance(skill_used_raw, list) or not all(isinstance(item, str) for item in skill_used_raw):
        raise ValueError("Generated JSON field 'skill_used' must be an array of strings.")
    if len(skill_used_raw) > 32 or any(len(item) > 512 for item in skill_used_raw):
        raise ValueError("Generated JSON field 'skill_used' exceeds the response contract limits.")
    notes = parsed.get("notes")
    if not isinstance(notes, str):
        raise ValueError("Generated JSON field 'notes' must be a string.")
    if len(notes) > MAX_GENERATION_NOTES_CHARS:
        raise ValueError(
            f"Generated notes exceed {MAX_GENERATION_NOTES_CHARS} characters: {len(notes)}."
        )
    return script_name, script, skill_used_raw, notes


def start_container(
    *,
    image: str,
    task_dir: Path,
    logs_dir: Path,
    task_id: str,
    task_text: str,
) -> tuple[str, CommandReport]:
    container = f"theking-sb-scripted-{safe_name(task_id)}-{uuid.uuid4().hex[:8]}"
    verifier_dir = task_dir / "verifier"
    logs_dir.mkdir(parents=True, exist_ok=True)
    (logs_dir / "verifier").mkdir(parents=True, exist_ok=True)
    report = run_command_with_tail(
        [
            "docker",
            "run",
            "-d",
            "--name",
            container,
            *docker_resource_args(task_text),
            "-v",
            f"{verifier_dir.resolve()}:/verifier:ro",
            "-v",
            f"{logs_dir.resolve()}:/logs",
            image,
            "sh",
            "-lc",
            "while true; do sleep 3600; done",
        ],
        timeout_sec=60,
    )
    return container, report


def inspect_container(container: str, *, workdir: str, timeout_sec: int) -> CommandReport:
    workdir_shell = shlex.quote(workdir)
    workdir_python = json.dumps(workdir)
    script = """
set +e
cd __WORKDIR_SHELL__ || exit 0
echo "## pwd"
pwd
echo "## file inventory"
if command -v find >/dev/null 2>&1; then
  find . -maxdepth 3 -type f -printf '%p\t%s bytes\n' 2>/dev/null | sort | head -200
fi
echo "## selected text heads"
for f in $(find . -maxdepth 2 -type f 2>/dev/null | sort | head -60); do
  case "$f" in
    *.txt|*.md|*.py|*.js|*.ts|*.java|*.yaml|*.yml|*.sh|*.toml|*.ini|*.conf)
      echo "--- $f"
      head -c 1200 "$f" 2>/dev/null
      echo
      ;;
  esac
done
if command -v python3 >/dev/null 2>&1; then
python3 - <<'PY'
from pathlib import Path
from collections import Counter
import csv
import json
root = Path(__WORKDIR_PYTHON__)

def short(value, limit=140):
    text = repr(value)
    if len(text) > limit:
        return text[:limit] + '...'
    return text

def dict_shape(obj, limit=20):
    return {str(k): type(v).__name__ for k, v in list(obj.items())[:limit]}

def coord_iter(obj, limit=5000):
    stack = [obj]
    seen = 0
    while stack and seen < limit:
        item = stack.pop()
        if isinstance(item, (list, tuple)):
            if len(item) >= 2 and all(isinstance(x, (int, float)) for x in item[:2]):
                seen += 1
                yield float(item[0]), float(item[1])
            else:
                stack.extend(reversed(item))

def geometry_bbox(geom):
    if not isinstance(geom, dict):
        return None
    coords = list(coord_iter(geom.get('coordinates')))
    if not coords:
        return None
    xs = [x for x, _ in coords]
    ys = [y for _, y in coords]
    return [round(min(xs), 6), round(min(ys), 6), round(max(xs), 6), round(max(ys), 6)]

def summarize_geojson(rel, data):
    features = data.get('features')
    if not isinstance(features, list):
        return False
    print(f'JSON {rel}: geojson-like features={len(features)} top_keys={list(data)[:20]} bbox={data.get("bbox")}')
    geom_types = Counter()
    property_keys = set()
    samples = []
    for feature in features[:20]:
        if not isinstance(feature, dict):
            continue
        geom = feature.get('geometry')
        if isinstance(geom, dict):
            geom_types[str(geom.get('type'))] += 1
        props = feature.get('properties')
        if isinstance(props, dict):
            property_keys.update(str(k) for k in props)
            if len(samples) < 3:
                samples.append({str(k): short(v, 80) for k, v in list(props.items())[:12]})
    print(f'  geometry_types_first20={dict(geom_types)}')
    print(f'  property_keys_sample={sorted(property_keys)[:40]}')
    for index, sample in enumerate(samples):
        print(f'  properties_sample_{index}={sample}')
    for index, feature in enumerate(features[:3]):
        if isinstance(feature, dict):
            print(f'  geometry_bbox_sample_{index}={geometry_bbox(feature.get("geometry"))}')
    return True

def summarize_json(rel, data):
    if isinstance(data, dict):
        if summarize_geojson(rel, data):
            return
        print(f'JSON {rel}: object keys={list(data)[:30]} shape={dict_shape(data)}')
        for key, value in list(data.items())[:10]:
            if isinstance(value, list):
                first = value[0] if value else None
                print(f'  {key}: list len={len(value)} first_type={type(first).__name__ if first is not None else None}')
                if isinstance(first, dict):
                    print(f'    first_shape={dict_shape(first)}')
            elif isinstance(value, dict):
                print(f'  {key}: object keys={list(value)[:20]} shape={dict_shape(value)}')
            else:
                print(f'  {key}: {type(value).__name__} sample={short(value)}')
    elif isinstance(data, list):
        first = data[0] if data else None
        print(f'JSON {rel}: list len={len(data)} first_type={type(first).__name__ if first is not None else None}')
        if isinstance(first, dict):
            keys = set()
            for item in data[:20]:
                if isinstance(item, dict):
                    keys.update(str(k) for k in item)
            print(f'  keys_sample={sorted(keys)[:40]} first_shape={dict_shape(first)}')
            for index, item in enumerate(data[:3]):
                sample = {str(k): short(v, 80) for k, v in list(item.items())[:12]}
                print(f'  row_sample_{index}={sample}')

def summarize_csv(rel, path):
    with path.open(newline='', encoding='utf-8', errors='replace') as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except Exception:
            dialect = csv.excel
        reader = csv.reader(handle, dialect)
        rows = []
        for _, row in zip(range(8), reader):
            rows.append(row[:20])
    header = rows[0] if rows else []
    print(f'CSV {rel}: header={header} sample_rows={rows[1:4]}')

print('## python data inspection')
for path in sorted(root.rglob('*')):
    if not path.is_file():
        continue
    rel = path.relative_to(root)
    suffix = path.suffix.lower()
    if suffix == '.json':
        try:
            data = json.loads(path.read_text(encoding='utf-8', errors='replace'))
            summarize_json(rel, data)
        except Exception as exc:
            print(f'JSON {rel}: unreadable {exc}')
    elif suffix == '.csv':
        try:
            summarize_csv(rel, path)
        except Exception as exc:
            print(f'CSV {rel}: metadata unavailable {exc}')
    elif suffix in {'.xlsx', '.xlsm'}:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=False)
            print(f'XLSX {rel}: sheets={wb.sheetnames}')
            for ws in wb.worksheets[:5]:
                print(f'  sheet={ws.title} max_row={ws.max_row} max_col={ws.max_column}')
                rows = []
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
                    values = [str(v) if v is not None else '' for v in row[:12]]
                    if any(values):
                        rows.append(values)
                for row in rows[:8]:
                    print('   ', row)
            wb.close()
        except Exception as exc:
            print(f'XLSX {rel}: metadata unavailable {exc}')
PY
fi
""".replace("__WORKDIR_SHELL__", workdir_shell).replace("__WORKDIR_PYTHON__", workdir_python)
    return run_command_with_tail(
        ["docker", "exec", container, "bash", "-lc", script],
        timeout_sec=timeout_sec,
    )


def build_prompt(
    *,
    task_id: str,
    task_text: str,
    arm: str,
    container_workdir: str,
    inspection: str,
    skill_context: list[dict[str, str]],
    previous_script: str | None = None,
    failure_context: str | None = None,
) -> str:
    if arm == "C1":
        skill_rule = (
            "Curated skills are provided below. Use them only when relevant. "
            "If you use a skill, list its top-level skill directory or SKILL.md path in skill_used."
        )
    else:
        skill_rule = "This is C0 no-skill. Do not use curated skill content. skill_used must be []."

    payload = {
        "task_id": task_id,
        "arm": arm,
        "harness_mode": "scripted_solver",
        "container_workdir": container_workdir,
        "skill_rule": skill_rule,
        "task_md": task_text,
        "inspection_report": inspection[:MAX_INSPECTION_CHARS],
        "skill_context": skill_context,
        "previous_script": previous_script,
        "failure_context": failure_context,
        "response_contract": {
            "script_path": "solve.sh or solution.py",
            "script": "complete executable script content",
            "skill_used": ["skill id or skill path; [] for C0"],
            "notes": "brief rationale",
        },
    }
    return (
        "You are generating a deterministic solve script for a SkillsBench task.\n"
        "Return strict JSON only. Do not include markdown fences or prose outside JSON.\n"
        f"Keep the complete script at or below {MAX_GENERATED_SCRIPT_CHARS} characters; prefer concise transformations over embedded data or long explanations.\n"
        f"The generated script will be copied to {container_workdir} inside the task Docker container and executed once.\n"
        f"The script must create or modify the required answer artifacts under {container_workdir}, unless task.md explicitly requires another absolute path.\n"
        "Do not rely on internet access. Do not ask follow-up questions.\n\n"
        + json.dumps(payload, ensure_ascii=False, indent=2)
    )


def copy_script_to_container(*, script_file: Path, container: str, workdir: str, script_name: str) -> CommandReport:
    target_dir = workdir.rstrip("/") or "/"
    return run_command_with_tail(
        ["docker", "cp", str(script_file.resolve()), f"{container}:{target_dir}/{script_name}"],
        timeout_sec=60,
    )


def run_script_in_container(*, container: str, workdir: str, script_name: str, timeout_sec: int) -> CommandReport:
    quoted_workdir = shlex.quote(workdir)
    quoted_script = shlex.quote(f"./{script_name}")
    if script_name.endswith(".py"):
        argv = ["docker", "exec", container, "bash", "-lc", f"cd {quoted_workdir} && python3 {quoted_script}"]
    else:
        argv = ["docker", "exec", container, "bash", "-lc", f"cd {quoted_workdir} && bash {quoted_script}"]
    return run_command_with_tail(argv, timeout_sec=timeout_sec)


def run_verifier_in_container(*, container: str, timeout_sec: int) -> CommandReport:
    return run_command_with_tail(
        ["docker", "exec", container, "bash", "/verifier/test.sh"],
        timeout_sec=timeout_sec,
    )


def cleanup_container(container: str) -> CommandReport:
    return run_command_with_tail(["docker", "rm", "-f", container], timeout_sec=60)


def failure_summary(*, commands: dict[str, CommandReport], reward: float | None) -> str:
    pieces = [f"reward={reward}"]
    for name in ("script", "verifier"):
        report = commands.get(name)
        if not report:
            continue
        pieces.append(
            json.dumps(
                {
                    "step": name,
                    "exit_code": report.exit_code,
                    "timed_out": report.timed_out,
                    "stdout_tail": report.stdout_tail[-3000:],
                    "stderr_tail": report.stderr_tail[-3000:],
                },
                ensure_ascii=False,
            )
        )
    return "\n".join(pieces)


def summarize(records: list[ScriptedRecord]) -> dict[str, Any]:
    by_key: dict[str, dict[str, Any]] = {}
    for record in records:
        key = f"{record.condition_id}:{record.arm}"
        slot = by_key.setdefault(
            key,
            {"n": 0, "passed": 0, "reward_observed": 0, "reward_sum": 0.0},
        )
        slot["n"] += 1
        slot["passed"] += int(record.passed)
        if record.reward is not None:
            slot["reward_observed"] += 1
            slot["reward_sum"] += record.reward
    for slot in by_key.values():
        observed = slot["reward_observed"]
        slot["mean_observed_reward"] = slot.pop("reward_sum") / observed if observed else None
    return {
        "record_count": len(records),
        "by_condition_arm": by_key,
    }


def append_record_jsonl(path: Path, record: ScriptedRecord) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(asdict(record), ensure_ascii=False) + "\n")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Run SkillsBench C0/C1 scripted-solver trials.")
    parser.add_argument("--task", action="append", help="Task id. May be repeated.")
    parser.add_argument("--limit-tasks", type=int, default=None)
    parser.add_argument("--adaptation-ready-first", action="store_true")
    parser.add_argument("--condition", action="append", help="Backend matrix condition id. May be repeated.")
    parser.add_argument("--matrix", type=Path, default=DEFAULT_MATRIX)
    parser.add_argument("--split-manifest", type=Path, default=DEFAULT_SPLIT_MANIFEST)
    parser.add_argument("--oracle-summary", type=Path, default=DEFAULT_ORACLE_SUMMARY)
    parser.add_argument("--run-id", default=datetime.now().strftime("%Y%m%d-%H%M%S"))
    parser.add_argument("--runs-root", type=Path, default=DEFAULT_RUNS_ROOT)
    parser.add_argument("--generation-timeout-sec", type=int, default=900)
    parser.add_argument("--script-timeout-sec", type=int, default=900)
    parser.add_argument("--verifier-timeout-sec", type=int, default=900)
    parser.add_argument("--build-timeout-sec", type=int, default=900)
    parser.add_argument("--inspection-timeout-sec", type=int, default=120)
    parser.add_argument(
        "--repair-iterations",
        type=int,
        default=0,
        help="Verifier-feedback retries. Keep 0 for evaluation; values >0 are adaptation-only.",
    )
    parser.add_argument("--trial-index", type=int, default=1)
    parser.add_argument("--keep-image", action="store_true")
    args = parser.parse_args(argv)

    run_start = time.monotonic()
    run_root = args.runs_root / args.run_id
    run_root.mkdir(parents=True, exist_ok=True)
    records_jsonl = run_root / "records.jsonl"
    records_jsonl.write_text("", encoding="utf-8")
    if args.adaptation_ready_first and not args.task:
        task_ids = load_adaptation_ready_tasks(args.split_manifest, args.oracle_summary)[:1]
    else:
        task_ids = args.task or list(DEFAULT_TASKS)
    if args.limit_tasks is not None:
        task_ids = task_ids[: args.limit_tasks]

    ready = ready_task_ids(args.oracle_summary)
    not_ready = [task_id for task_id in task_ids if task_id not in ready]
    if not_ready:
        raise SystemExit(f"Tasks are not in executable readiness passed set: {', '.join(not_ready)}")

    conditions = load_conditions(args.matrix, args.condition)
    backend_versions = {
        condition["id"]: probe_backend_version(condition)
        for condition in conditions
    }
    records: list[ScriptedRecord] = []
    dockerd_proc = None
    try:
        dockerd_proc = ensure_docker(run_root)
        for task_id in task_ids:
            task_dir = TASKS_ROOT / task_id
            task_source = read_task_prompt(task_dir)
            task_text = task_instruction_body(task_source)
            timeouts = task_timeouts(
                task_source,
                build_default=args.build_timeout_sec,
                script_default=args.script_timeout_sec,
                verifier_default=args.verifier_timeout_sec,
            )
            record_timeouts = {
                "generation_timeout_sec": args.generation_timeout_sec,
                "inspection_timeout_sec": args.inspection_timeout_sec,
                **timeouts,
            }
            image = f"theking-skillsbench-scripted-{safe_name(task_id)}:latest"
            with tempfile.TemporaryDirectory(prefix=f"build-{safe_name(task_id)}-", dir=run_root) as tmp:
                build_context = prepare_skill_free_build_context(
                    task_dir / "environment",
                    Path(tmp) / "environment",
                )
                build = run_command_with_tail(
                    ["docker", "build", "-t", image, str(build_context)],
                    timeout_sec=timeouts["build_timeout_sec"],
                )
            if build.exit_code != 0:
                for condition in conditions:
                    for arm in ("C0", "C1"):
                        record = ScriptedRecord(
                            task_id=task_id,
                            condition_id=condition["id"],
                            arm=arm,
                            harness_mode="scripted_repair" if args.repair_iterations else "scripted_solver",
                            model_id=condition["model_id"],
                            backend=condition["backend"],
                            effort=condition["effort"],
                            runtime_effort=condition.get("runtime_effort", condition["effort"]),
                            status="build_failed",
                            passed=False,
                            trial_index=args.trial_index,
                            trial_id=f"{args.run_id}:{task_id}:{condition['id']}:{arm}:trial-{args.trial_index}",
                            wall_time_sec=build.duration_sec,
                            timeouts=record_timeouts,
                            commands={"build": build},
                        )
                        records.append(record)
                        append_record_jsonl(records_jsonl, record)
                continue

            container_workdir, workdir_report = detect_image_workdir(image)
            image_id_report = run_command_with_tail(
                ["docker", "image", "inspect", "-f", "{{.Id}}", image],
                timeout_sec=30,
            )
            try:
                for condition in conditions:
                    for arm in ("C0", "C1"):
                        trial_start = time.monotonic()
                        condition_dir = run_root / "tasks" / task_id / condition["id"] / arm
                        generated_dir = condition_dir / "generated"
                        logs_dir = condition_dir / "logs"
                        if condition_dir.exists():
                            shutil.rmtree(condition_dir)
                        generated_dir.mkdir(parents=True, exist_ok=True)
                        logs_dir.mkdir(parents=True, exist_ok=True)

                        commands: dict[str, CommandReport] = {
                            "build": build,
                            "image_workdir": workdir_report,
                            "image_id": image_id_report,
                            "backend_version": backend_versions[condition["id"]],
                        }
                        notes: list[str] = []
                        usage_attempts: list[dict[str, Any]] = []
                        skill_context = collect_skill_context(task_dir, arm=arm)
                        skill_context_report = describe_skill_context(
                            task_dir,
                            arm=arm,
                            injected=skill_context,
                        )
                        previous_script: str | None = None
                        failure_context: str | None = None
                        final_script_name: str | None = None
                        final_skill_used: list[str] = []
                        final_reward: float | None = None
                        final_prompt_sha256: str | None = None
                        status = "generation_failed"
                        passed = False
                        repairs_used = 0

                        for attempt in range(args.repair_iterations + 1):
                            container, start_report = start_container(
                                image=image,
                                task_dir=task_dir,
                                logs_dir=logs_dir,
                                task_id=task_id,
                                task_text=task_source,
                            )
                            commands[f"container_start_{attempt}"] = start_report
                            if start_report.exit_code != 0:
                                status = "container_start_failed"
                                break
                            try:
                                inspection_report = inspect_container(
                                    container,
                                    workdir=container_workdir,
                                    timeout_sec=args.inspection_timeout_sec,
                                )
                                commands[f"inspection_{attempt}"] = inspection_report
                                prompt = build_prompt(
                                    task_id=task_id,
                                    task_text=task_text,
                                    arm=arm,
                                    container_workdir=container_workdir,
                                    inspection=inspection_report.stdout_tail,
                                    skill_context=skill_context,
                                    previous_script=previous_script,
                                    failure_context=failure_context,
                                )
                                final_prompt_sha256 = hashlib.sha256(prompt.encode("utf-8")).hexdigest()
                                (generated_dir / f"attempt-{attempt}-prompt.txt").write_text(
                                    prompt,
                                    encoding="utf-8",
                                )
                                generation = call_generation_model(
                                    condition,
                                    prompt,
                                    cwd=condition_dir,
                                    timeout_sec=args.generation_timeout_sec,
                                )
                                commands[f"generation_{attempt}"] = generation
                                raw_generation_path = generated_dir / f"attempt-{attempt}-generation-raw.json"
                                raw_generation_path.write_text(generation.stdout_tail, encoding="utf-8")
                                usage = extract_account_usage(generation)
                                if usage:
                                    usage_attempts.append(usage)
                                if generation.exit_code != 0:
                                    status = "generation_timeout" if generation.timed_out else "generation_failed"
                                    trim_generation_trace(generation)
                                    break
                                try:
                                    script_name, script, skill_used, note = parse_generated_script(generation)
                                except Exception as exc:
                                    notes.append(f"parse_error_attempt_{attempt}:{exc}")
                                    status = generation_parse_failure_status(generation.stdout_tail)
                                    trim_generation_trace(generation)
                                    break
                                trim_generation_trace(generation)
                                if note:
                                    notes.append(f"model_note_attempt_{attempt}:{note}")
                                final_script_name = script_name
                                final_skill_used = skill_used
                                script_file = generated_dir / f"attempt-{attempt}-{script_name}"
                                script_file.write_text(script, encoding="utf-8")
                                copy_report = copy_script_to_container(
                                    script_file=script_file,
                                    container=container,
                                    workdir=container_workdir,
                                    script_name=script_name,
                                )
                                commands[f"copy_script_{attempt}"] = copy_report
                                if copy_report.exit_code != 0:
                                    status = "script_copy_failed"
                                    break
                                script_report = run_script_in_container(
                                    container=container,
                                    workdir=container_workdir,
                                    script_name=script_name,
                                    timeout_sec=timeouts["script_execution_timeout_sec"],
                                )
                                commands[f"script_{attempt}"] = script_report
                                verifier_report = run_verifier_in_container(
                                    container=container,
                                    timeout_sec=timeouts["verifier_timeout_sec"],
                                )
                                commands[f"verifier_{attempt}"] = verifier_report
                                final_reward = read_reward(logs_dir)
                                status, passed = verifier_outcome_status(verifier_report, final_reward)
                                if passed:
                                    break
                                if attempt < args.repair_iterations:
                                    repairs_used += 1
                                    previous_script = script
                                    failure_context = failure_summary(
                                        commands={
                                            "script": script_report,
                                            "verifier": verifier_report,
                                        },
                                        reward=final_reward,
                                    )
                            finally:
                                commands[f"container_cleanup_{attempt}"] = cleanup_container(container)
                            if passed or attempt >= args.repair_iterations:
                                break

                        record = ScriptedRecord(
                            task_id=task_id,
                            condition_id=condition["id"],
                            arm=arm,
                            harness_mode="scripted_repair" if args.repair_iterations else "scripted_solver",
                            model_id=condition["model_id"],
                            backend=condition["backend"],
                            effort=condition["effort"],
                            runtime_effort=condition.get("runtime_effort", condition["effort"]),
                            status=status,
                            passed=passed,
                            trial_index=args.trial_index,
                            trial_id=f"{args.run_id}:{task_id}:{condition['id']}:{arm}:trial-{args.trial_index}",
                            reward=final_reward,
                            wall_time_sec=round(time.monotonic() - trial_start, 3),
                            container_workdir=container_workdir,
                            script_path=final_script_name,
                            skill_used=final_skill_used,
                            task_instruction_sha256=hashlib.sha256(task_text.encode("utf-8")).hexdigest(),
                            prompt_sha256=final_prompt_sha256,
                            skill_context_report=skill_context_report,
                            repair_iterations_used=repairs_used,
                            account_usage={"generation_attempts": usage_attempts},
                            timeouts=record_timeouts,
                            logs_dir=str(logs_dir),
                            commands=commands,
                            notes=notes,
                        )
                        records.append(record)
                        append_record_jsonl(records_jsonl, record)
                        print(json.dumps(asdict(record), ensure_ascii=False), flush=True)
            finally:
                if not args.keep_image:
                    run_command_with_tail(["docker", "rmi", "-f", image], timeout_sec=120)
    finally:
        stop_process(dockerd_proc)

    output = {
        "run_id": args.run_id,
        "harness_mode": "scripted_repair" if args.repair_iterations else "scripted_solver",
        "benchmark_eligibility": {
            "skillsbench_paper_c0_c1": False,
            "reason": "bounded scripted harness; C1 prompt injection is not native full-bundle agent interaction",
        },
        "backend_contract": {
            "type": "B_cli",
            "auth_mode": "user_owned_account",
            "api_keys_required": False,
            "credentials_forwarded_to_container": False,
        },
        "trial_control": {
            "trial_index": args.trial_index,
            "seed_control": "provider_cli_unavailable",
            "temperature_control": "provider_default_or_unset",
            "paper_required_trials_per_cell": 3,
        },
        "response_contract": {
            "version": RESPONSE_CONTRACT_VERSION,
            "provider_structured_output": "claude --json-schema",
            "max_script_chars": MAX_GENERATED_SCRIPT_CHARS,
            "max_notes_chars": MAX_GENERATION_NOTES_CHARS,
        },
        "task_ids": task_ids,
        "condition_ids": [condition["id"] for condition in conditions],
        "backend_versions": {
            condition_id: asdict(report)
            for condition_id, report in backend_versions.items()
        },
        "timeouts": {
            "generation_timeout_sec": args.generation_timeout_sec,
            "script_execution_timeout_sec": args.script_timeout_sec,
            "verifier_timeout_sec": args.verifier_timeout_sec,
            "build_timeout_sec": args.build_timeout_sec,
            "inspection_timeout_sec": args.inspection_timeout_sec,
            "task_frontmatter_budget_overrides": True,
            "task_frontmatter_exposed_to_model": False,
        },
        "repair_iterations": args.repair_iterations,
        "repair_policy": "evaluation_zero_retry" if args.repair_iterations == 0 else "adaptation_only_verifier_feedback",
        "wall_time_sec": round(time.monotonic() - run_start, 3),
        "records": [asdict(record) for record in records],
        "summary": summarize(records),
    }
    (run_root / "summary.json").write_text(json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(output["summary"], ensure_ascii=False), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
